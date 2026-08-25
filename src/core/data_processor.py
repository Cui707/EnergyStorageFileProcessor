"""
Core data processor for single file processing.
"""

import os
import time
import pandas as pd
from typing import Dict, Any, List, Optional
from ..config import Config
from ..models.data_models import BatteryData, ProcessedData, ProcessingResult
from ..readers.factory import ReaderFactory
from ..processors.extractor import DataExtractor
from ..processors.calculator import DataCalculator
from ..processors.validator import DataValidator
from ..utils.logger import Logger
from ..utils.file_utils import FileUtils


class DataProcessor:
    """Main data processor for single file processing."""
    
    def __init__(self, config: Config):
        """
        Initialize the data processor.
        
        Args:
            config: Configuration instance
        """
        self.config = config
        self.logger = Logger("DataProcessor")
        
        # Initialize components
        self.reader_factory = ReaderFactory()
        self.extractor = DataExtractor(config)
        self.calculator = DataCalculator(config)
        self.validator = DataValidator(config)
        
        # Get output settings
        self.output_format = config.get_output_format()
        self.output_filename = config.get_output_filename()
        self.highlight_calculated = config.should_highlight_calculated()
    
    def process_file(self, input_file: str, output_dir: str, 
                    custom_output_name: Optional[str] = None) -> ProcessingResult:
        """
        Process a single data file.
        
        Args:
            input_file: Path to input file
            output_dir: Directory for output file
            custom_output_name: Custom output filename (optional)
            
        Returns:
            ProcessingResult object
        """
        start_time = time.time()
        
        try:
            # Validate input file
            if not os.path.exists(input_file):
                return ProcessingResult(
                    input_file=input_file,
                    output_file="",
                    success=False,
                    error_message="Input file not found"
                )
            
            # Create output directory
            FileUtils.ensure_directory(output_dir)
            
            # Detect file type and create reader
            file_type = self.reader_factory.auto_detect_reader(input_file, 
                                                             self.config.get_reader_settings())
            reader = self.reader_factory.create_reader(file_type, 
                                                     self.config.get_reader_settings())
            
            # Read data
            self.logger.log_processing_start(input_file, 1)
            raw_data = reader.read(input_file)
            
            # Validate raw data
            is_valid, validation_errors = self.validator.validate_dataframe(raw_data)
            if not is_valid:
                return ProcessingResult(
                    input_file=input_file,
                    output_file="",
                    success=False,
                    error_message=f"Data validation failed: {', '.join(validation_errors)}"
                )
            
            # Extract cluster ID from filename
            cluster_id = reader.extract_cluster_id(input_file)
            
            # Extract battery data
            battery_data = self.extractor.extract_battery_data(raw_data, cluster_id)
            
            # Validate battery data
            is_valid, validation_errors = self.validator.validate_battery_data(battery_data)
            if not is_valid:
                return ProcessingResult(
                    input_file=input_file,
                    output_file="",
                    success=False,
                    error_message=f"Battery data validation failed: {', '.join(validation_errors)}"
                )
            
            # Calculate system statistics
            processed_data = self.calculator.calculate_system_statistics(battery_data)
            
            # Create output DataFrame
            output_df = processed_data.to_dataframe()
            
            # Add calculated columns
            output_df = self.calculator.add_calculated_columns(output_df, processed_data)
            
            # Generate output filename
            if custom_output_name:
                output_filename = custom_output_name
            else:
                base_name = os.path.splitext(os.path.basename(input_file))[0]
                output_filename = f"{base_name}_{self.output_filename}.{self.output_format}"
            
            output_file = os.path.join(output_dir, output_filename)
            
            # Save output file
            self._save_output_file(output_df, output_file)
            
            processing_time = time.time() - start_time
            records_processed = len(output_df)
            
            self.logger.log_processing_complete(input_file, output_file, 
                                             processing_time, records_processed)
            
            return ProcessingResult(
                input_file=input_file,
                output_file=output_file,
                success=True,
                processing_time=processing_time,
                records_processed=records_processed
            )
            
        except Exception as e:
            processing_time = time.time() - start_time
            error_msg = f"Error processing file: {str(e)}"
            
            self.logger.log_processing_error(input_file, error_msg)
            
            return ProcessingResult(
                input_file=input_file,
                output_file="",
                success=False,
                error_message=error_msg,
                processing_time=processing_time
            )
    
    def _save_output_file(self, df: pd.DataFrame, output_file: str):
        """
        Save output file based on format.
        
        Args:
            df: DataFrame to save
            output_file: Output file path
        """
        if self.output_format.lower() == 'excel':
            self._save_excel_file(df, output_file)
        elif self.output_format.lower() == 'csv':
            self._save_csv_file(df, output_file)
        else:
            raise ValueError(f"Unsupported output format: {self.output_format}")
    
    def _save_excel_file(self, df: pd.DataFrame, output_file: str):
        """Save DataFrame to Excel file."""
        try:
            # Save to Excel
            df.to_excel(output_file, index=False)
            
            # Apply formatting if needed
            if self.highlight_calculated:
                self._apply_excel_formatting(output_file)
                
        except Exception as e:
            raise Exception(f"Failed to save Excel file: {str(e)}")
    
    def _save_csv_file(self, df: pd.DataFrame, output_file: str):
        """Save DataFrame to CSV file."""
        try:
            df.to_csv(output_file, index=False)
        except Exception as e:
            raise Exception(f"Failed to save CSV file: {str(e)}")
    
    def _apply_excel_formatting(self, output_file: str):
        """Apply Excel formatting to highlight calculated columns."""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font
            
            # Load workbook
            wb = load_workbook(output_file)
            ws = wb.active
            
            # Define calculated columns to highlight
            calculated_columns = [
                'sysMaxU', 'sysMinU', 'MaxDiff', 
                'sysMaxT', 'DayTotalChargeKwh', 'DayTotalDischargeKwh'
            ]
            
            # Set red font for calculated columns
            red_font = Font(color="FF0000")
            
            for col_name in calculated_columns:
                if col_name in ws[1]:  # Check if column exists
                    col_idx = ws[1].index(col_name) + 1  # Convert to 1-based index
                    for row in range(2, len(ws) + 1):  # Skip header row
                        cell = ws.cell(row=row, column=col_idx)
                        cell.font = red_font
            
            # Save workbook
            wb.save(output_file)
            
        except Exception as e:
            self.logger.warning(f"Failed to apply Excel formatting: {str(e)}")
    
    def get_supported_formats(self) -> List[str]:
        """Get list of supported input formats."""
        return self.reader_factory.get_supported_types()
    
    def validate_input_file(self, file_path: str) -> bool:
        """
        Validate if input file can be processed.
        
        Args:
            file_path: Path to input file
            
        Returns:
            True if file can be processed
        """
        try:
            if not os.path.exists(file_path):
                return False
            
            # Check file extension
            extension = os.path.splitext(file_path)[1].lower()
            if extension not in ['.csv', '.xlsx', '.xls']:
                return False
            
            # Try to create reader
            reader = self.reader_factory.auto_detect_reader(
                file_path, 
                self.config.get_reader_settings()
            )
            
            return True
            
        except Exception:
            return False