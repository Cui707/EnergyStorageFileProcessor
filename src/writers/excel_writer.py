"""
Excel writer for data files.
"""

import pandas as pd
from typing import Dict, Any, Optional
from .base import BaseWriter


class ExcelWriter(BaseWriter):
    """Excel file writer for data."""
    
    def __init__(self, config: Dict[str, Any]):
        """
        Initialize the Excel writer.
        
        Args:
            config: Configuration dictionary
        """
        super().__init__(config)
        self.highlight_calculated = config.get('highlight_calculated', True)
    
    def write(self, data: pd.DataFrame, file_path: str) -> bool:
        """
        Write data to Excel file.
        
        Args:
            data: DataFrame to write
            file_path: Path to output file
            
        Returns:
            True if write successful
        """
        try:
            # Ensure output directory exists
            if not self.ensure_directory(file_path):
                raise Exception(f"Cannot create output directory: {Path(file_path).parent}")
            
            # Write to Excel
            data.to_excel(file_path, index=False)
            
            # Apply formatting if needed
            if self.highlight_calculated:
                self._apply_excel_formatting(file_path)
            
            return True
            
        except Exception as e:
            print(f"Failed to write Excel file {file_path}: {e}")
            return False
    
    def _apply_excel_formatting(self, file_path: str):
        """
        Apply Excel formatting to highlight calculated columns.
        
        Args:
            file_path: Path to Excel file
        """
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font
            
            # Load workbook
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Define calculated columns to highlight
            calculated_columns = [
                'sysMaxU', 'sysMinU', 'MaxDiff', 
                'sysMaxT', 'DayTotalChargeKwh', 'DayTotalDischargeKwh'
            ]
            
            # Set red font for calculated columns
            red_font = Font(color="FF0000")
            
            # Find column indices for calculated columns
            header_row = ws[1]  # First row is header
            for col_name in calculated_columns:
                if col_name in header_row:
                    col_idx = list(header_row).index(col_name) + 1  # Convert to 1-based index
                    for row in range(2, len(ws) + 1):  # Skip header row
                        cell = ws.cell(row=row, column=col_idx)
                        cell.font = red_font
            
            # Save workbook
            wb.save(file_path)
            
        except Exception as e:
            print(f"Failed to apply Excel formatting: {e}")
    
    def get_supported_extensions(self) -> list:
        """Get supported file extensions."""
        return ['.xlsx', '.xls']